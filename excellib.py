import sys, os, threading
from xlwt import *
import xlrd
import openpyxl

def save_xlsx(file_path, data):
	book = openpyxl.Workbook()
	sheet = book.active
	sheet.title = 'Sheet1'
		
	rowindex = 1
	for i, r in enumerate (data):
		headers = []
		values = []
		for j, col in enumerate(r):
			if j % 2 == 0:
				headers.append(col)
			else:			
				if isinstance(col, basestring):
					col = col.strip()			
				values.append(col)
		if i==0:
			#write headers
			
			for colindex, h in enumerate(headers):
				sheet.cell(row=1,column=colindex+1).value = h

		rowindex += 1
		for colindex, value in enumerate(values):
			sheet.cell(row=rowindex,column=colindex+1).value = value if value is not None else ''

	book.save(file_path)		

def save_xls(file_path, data):
	
	book = Workbook(encoding="utf-8")
	
	sheet = book.add_sheet("Sheet1")
	style = XFStyle()
	style.num_format_str = '0.00'
	rowindex = 0
	for i, r in enumerate (data):
		headers = []
		values = []
		for j, col in enumerate(r):
			if j % 2 == 0:
				headers.append(col)
			else:			
				if isinstance(col, basestring):
					col = col.strip()			
				values.append(col)
		if i==0:
			#write headers
			for colindex, h in enumerate(headers):
				sheet.write(0,colindex,h)

		rowindex += 1
		for colindex, value in enumerate(values):
			if value is not None:			
				sheet.write(rowindex,colindex,value, style)	
			else:				
				sheet.write(rowindex,colindex,'', style)	

	book.save(file_path)			

def csv_data_to_xls(file_path, data):
	
	book = Workbook(encoding="utf-8")
	
	sheet = book.add_sheet("sheet1")
	style = XFStyle()
	style.num_format_str = '0.00'
	rowindex = -1
	for r in data:
		rowindex += 1
		
		for colindex, value in enumerate(r):
			sheet.write(rowindex,colindex,value, style)	

	book.save(file_path)			

def read_xlsx_sheet(file_path, index=0):
	wb = openpyxl.load_workbook(file_path)
	
	sheet = wb.worksheets[index]	
	data = []
	for row in sheet.rows:
		r = []
		for cell in row:
			r.append(cell.value)
		data.append(r)	

	return data	

def read_sheet(file_path, return_type='list', index=0):
	"""
	return_type: list, dict
	"""
	data = []
	if file_path.lower().endswith('.xlsx'):
		data = read_xlsx_sheet(file_path)
	else:	
		book = xlrd.open_workbook(file_path)	
		sheet1 = book.sheet_by_index(index)

		for i in range(sheet1.nrows):
			r = sheet1.row_values(i)		
			data.append(r)
	
	fields =[]

	for field in data[0]:
		if not field: break #every fields after the first None field are ignored
		fields.append(field)

	
	if return_type == 'list':
		if len(fields)<len(data[0]):
			cleaneddata = [r[0:len(fields)] for r in data]
			return cleaneddata
		else:	
			return data

	
	rs = []
	rowindex = 1
	for r in data[1:]:
		rowindex += 1
		if len(r) != len(data[0]):
			raise Exception("Inconsistent row length at row#: %s" % rowindex)

		row = {}	
		for i, value in enumerate(r):
			if i== len(fields): break			
			row.update({fields[i]: value})
		rs.append(row)	
		
	return rs	


if __name__ == '__main__':
	print 'test'
	data = [['Cung', 'Nguyen']]
	save_xlsx('/Users/cung/test.xlsx', data)




